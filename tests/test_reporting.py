import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.models import (
    CommentRecord,
    CommentWatch,
    ContentRecord,
    DanmakuRecord,
    DanmakuWatch,
    MonitorTarget,
)
from app.reporting import (
    build_comments_report,
    build_contents_report,
    build_danmaku_report,
    build_danmaku_watches_report,
    build_monitor_report,
    build_share_history_report,
    build_targets_report,
    build_watches_report,
)


class MonitorReportTests(unittest.TestCase):
    def test_report_contains_summary_and_monitoring_details(self):
        captured_at = datetime(2026, 8, 4, 10, 30)
        target = MonitorTarget(
            id=1,
            platform="douyin",
            sec_uid="target-sec-uid",
            nickname="Sample creator",
            alias="Sample",
            group_name="priority",
            tags='["hot"]',
            created_at=captured_at,
        )
        content = ContentRecord(
            id=2,
            platform="douyin",
            target_id=1,
            aweme_id="aweme-1",
            desc="=SUM(1, 1)",
            create_time=1_700_000_000,
            like_count=12,
            comment_count=3,
            download_status="done",
            created_at=captured_at,
        )
        comment_watch = CommentWatch(
            id=3,
            platform="douyin",
            aweme_id="aweme-1",
            title="Sample work",
            alias="Comments",
        )
        comment = CommentRecord(
            id=4,
            platform="douyin",
            watch_id=3,
            aweme_id="aweme-1",
            comment_id="comment-1",
            text="+engagement",
            created_at=captured_at,
        )
        danmaku_watch = DanmakuWatch(
            id=5,
            platform="douyin",
            aweme_id="aweme-1",
            title="Sample work danmaku",
        )
        danmaku = DanmakuRecord(
            id=6,
            platform="douyin",
            watch_id=5,
            aweme_id="aweme-1",
            danmaku_id="danmaku-1",
            text="@viewer",
            created_at=captured_at,
        )

        payload = build_monitor_report(
            platform="douyin",
            period_start=None,
            period_end=None,
            targets=[target],
            contents=[content],
            watches=[comment_watch],
            comments=[comment],
            danmaku_watches=[danmaku_watch],
            danmaku=[danmaku],
            generated_at=captured_at,
        )

        workbook = load_workbook(BytesIO(payload), data_only=False)
        self.assertEqual(len(workbook.worksheets), 5)
        self.assertEqual(workbook.worksheets[0]["B10"].value, 1)
        self.assertEqual(workbook.worksheets[0]["B11"].value, 1)
        self.assertEqual(workbook.worksheets[0]["B12"].value, 1)
        self.assertEqual(workbook.worksheets[0]["B13"].value, 1)
        self.assertEqual(workbook.worksheets[2]["D2"].value, "aweme-1")
        self.assertEqual(workbook.worksheets[2]["E2"].value, "'=SUM(1, 1)")
        self.assertEqual(workbook.worksheets[3]["F2"].value, "'+engagement")
        self.assertEqual(workbook.worksheets[4]["F2"].value, "'@viewer")

    def test_report_can_be_built_with_empty_data(self):
        payload = build_monitor_report(
            platform="xhs",
            period_start=None,
            period_end=None,
            targets=[],
            contents=[],
            watches=[],
            comments=[],
            danmaku_watches=[],
            danmaku=[],
        )
        workbook = load_workbook(BytesIO(payload), read_only=True)
        self.assertEqual(len(workbook.worksheets), 5)
        self.assertGreater(len(payload), 1000)

    def test_each_module_report_contains_only_that_module(self):
        target = MonitorTarget(id=1, platform="douyin", nickname="target")
        content = ContentRecord(id=2, platform="douyin", target_id=1, aweme_id="a")
        watch = CommentWatch(id=3, platform="douyin", aweme_id="a", title="watch")
        comment = CommentRecord(
            id=4, platform="douyin", watch_id=3, aweme_id="a", comment_id="c"
        )
        danmaku_watch = DanmakuWatch(
            id=5, platform="douyin", aweme_id="a", title="danmaku watch"
        )
        danmaku = DanmakuRecord(
            id=6, platform="douyin", watch_id=5, aweme_id="a", danmaku_id="d"
        )
        reports = [
            build_targets_report([target], [content]),
            build_contents_report([content], [target]),
            build_watches_report([watch]),
            build_comments_report([comment], [watch]),
            build_danmaku_watches_report([danmaku_watch]),
            build_danmaku_report([danmaku], [danmaku_watch]),
        ]
        for payload in reports:
            workbook = load_workbook(BytesIO(payload), read_only=True)
            self.assertEqual(len(workbook.worksheets), 2)

    def test_module_report_keeps_filter_context_and_presentation_formatting(self):
        captured_at = datetime(2026, 8, 4, 10, 30)
        target = MonitorTarget(id=1, platform="douyin", nickname="target")
        content = ContentRecord(
            id=2,
            platform="douyin",
            target_id=1,
            aweme_id="a",
            desc="a long description",
            like_count=1234,
            download_status="done",
            cover_url="https://example.test/cover.jpg",
            local_path="data/media/sample.mp4",
            created_at=captured_at,
        )
        workbook = load_workbook(BytesIO(build_contents_report(
            [content],
            [target],
            filters=[("scope", "filtered"), ("platform", "douyin")],
            generated_at=captured_at,
        )))
        summary, data = workbook.worksheets
        self.assertEqual(summary["B4"].value, 1)
        self.assertEqual(summary.cell(summary.max_row, 2).hyperlink.location, "'作品数据'!A1")
        self.assertEqual(data.freeze_panes, "E2")
        self.assertEqual(data["I2"].number_format, "#,##0")
        local_path = str(Path("data/media/sample.mp4").resolve())
        self.assertEqual(
            data["N2"].value,
            f'=HYPERLINK("{local_path}","{local_path}")',
        )
        self.assertIsNone(data["N2"].hyperlink)
        self.assertEqual(data["O2"].hyperlink.target, "https://example.test/cover.jpg")
        self.assertEqual(data["M2"].fill.fgColor.rgb, "00ECFDF5")

    def test_share_history_report_contains_download_fields(self):
        captured_at = datetime(2026, 8, 4, 10, 30)
        workbook = load_workbook(BytesIO(build_share_history_report([
            {
                "id": 7,
                "platform": "douyin",
                "item_id": "aweme-7",
                "title": "Sample download",
                "author": "Sample author",
                "media_type": "video",
                "media_count": 1,
                "create_time": 1_700_000_000,
                "created_at": captured_at,
                "like_count": 12,
                "comment_count": 3,
                "duration": 42,
                "quality": "1080P",
                "status": "done",
                "output_dir": "data/media/share",
                "files": [{"role": "media", "path": "data/media/share/sample.mp4"}],
                "source_url": "https://example.test/share/7",
            }
        ], generated_at=captured_at)))
        summary, data = workbook.worksheets
        self.assertEqual(summary["B4"].value, 1)
        self.assertEqual(data["A2"].value, 7)
        self.assertEqual(data["N2"].value, "done")
        self.assertEqual(data["I2"].value, captured_at)
        self.assertTrue(str(data["O2"].value).startswith("=HYPERLINK("))
        self.assertTrue(str(data["P2"].value).startswith("=HYPERLINK("))
        self.assertEqual(data["Q2"].hyperlink.target, "https://example.test/share/7")


if __name__ == "__main__":
    unittest.main()
