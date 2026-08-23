import asyncio
import json
import tempfile
import unittest
from contextlib import asynccontextmanager
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock

from openpyxl import load_workbook
from fastapi import HTTPException
from sqlmodel import select

from app import db
from app.browser.fetcher import (
    _douyin_search_item_matches,
    _sort_douyin_search_items,
    _douyin_search_needs_verification,
    douyin_search_empty_error,
    douyin_search_exception_error,
    extract_search_awemes,
    fetch_douyin_search,
)


def _req():
    """直调端点函数的假 request(测试旁路环境下视为管理员)。"""
    return SimpleNamespace(state=SimpleNamespace(user=SimpleNamespace(id=1, is_superuser=True, role="admin")))


from app.config import Config
from app.engine.collection import KeywordCollector
from app.main import (
    KeywordCollectionIn,
    _collection_content_dict,
    _collection_error_for_display,
    _collection_keywords,
    create_keyword_collection,
    keyword_collection_content_media,
    keyword_collection_local_media,
    update_keyword_collection,
)
from app.models import (
    DouyinAccount,
    KeywordCollectionComment,
    KeywordCollectionContent,
    KeywordCollectionJob,
)
from app.reporting import build_keyword_collection_report


class SearchExtractionTests(unittest.TestCase):
    def test_douyin_verify_check_payload_is_detected(self):
        self.assertTrue(_douyin_search_needs_verification({
            "data": [],
            "search_nil_info": {
                "search_nil_type": "verify_check",
                "search_nil_item": "verify_check",
            },
        }))
        self.assertFalse(_douyin_search_needs_verification({
            "data": [{"aweme_info": {"aweme_id": "1"}}],
        }))

    def test_collection_error_display_shortens_closed_browser_trace(self):
        message = _collection_error_for_display(
            "ai短剧: 打开抖音搜索页失败: TargetClosedError('Target page, "
            "context or browser has been closed')")
        self.assertEqual(
            message,
            "ai短剧：采集窗口已关闭，请点击“续跑”并保持窗口开启",
        )
        self.assertNotIn("TargetClosedError", message)

    def test_captcha_page_has_actionable_error(self):
        message = douyin_search_empty_error(
            "验证码中间页", "", "https://www.douyin.com/search/example", [])
        self.assertIn("打开浏览器", message)
        self.assertIn("续跑", message)

    def test_login_modal_is_not_misclassified_as_security_captcha(self):
        message = douyin_search_empty_error(
            "抖音搜索", "扫码登录 验证码登录 登录后即可搜索更多精彩视频",
            "https://www.douyin.com/search/example", [])
        self.assertEqual(message, "抖音登录态已失效；请重新扫码登录后再续跑")

    def test_closed_visible_window_has_short_actionable_error(self):
        message = douyin_search_exception_error(
            RuntimeError("TargetClosedError: Target page, context or browser has been closed"))
        self.assertEqual(
            message,
            "抖音采集窗口被关闭；请续跑任务，并在任务结束前保持窗口打开",
        )

    def test_keyword_input_is_split_normalized_and_deduplicated(self):
        self.assertEqual(
            _collection_keywords(["露营, 骑行", "露营\n穿搭", "骑行"]),
            ["露营", "骑行", "穿搭"],
        )

    def test_extracts_known_search_wrappers_and_deduplicates(self):
        first = {
            "aweme_id": "a-1",
            "desc": "first",
            "video": {"play_addr": {"url_list": ["https://media.test/a.mp4"]}},
        }
        duplicate = {**first, "desc": "duplicate"}
        second = {
            "aweme_id": "a-2",
            "images": [{"url_list": ["https://media.test/b.jpg"]}],
        }
        payload = {
            "data": [
                {"aweme_info": first},
                {"search_result": {"aweme_mix_info": duplicate}},
                {"card": {"items": [second]}},
                {"aweme_info": {"aweme_id": "user-card-without-media"}},
            ]
        }

        result = extract_search_awemes(payload)

        self.assertEqual([row["aweme_id"] for row in result], ["a-1", "a-2"])
        self.assertEqual(result[0]["desc"], "first")

    def test_search_filters_and_sorting_are_applied_deterministically(self):
        rows = [
            {"aweme_id": "old", "create_time": 100, "video": {"play_addr": {}},
             "statistics": {"digg_count": 100, "comment_count": 8}},
            {"aweme_id": "new", "create_time": 190, "images": [{}],
             "statistics": {"digg_count": 20, "comment_count": 3}},
            {"aweme_id": "hot", "create_time": 180, "video": {"play_addr": {}},
             "statistics": {"digg_count": 500, "comment_count": 20}},
        ]

        self.assertFalse(_douyin_search_item_matches(
            rows[0], content_type="video", publish_time="day",
            min_likes=50, min_comments=5, now=100_000,
        ))
        self.assertTrue(_douyin_search_item_matches(
            rows[2], content_type="video", min_likes=100, min_comments=10,
        ))
        self.assertEqual(
            [row["aweme_id"] for row in _sort_douyin_search_items(rows, "latest")],
            ["new", "hot", "old"],
        )
        self.assertEqual(
            [row["aweme_id"] for row in _sort_douyin_search_items(rows, "most_liked")],
            ["hot", "old", "new"],
        )


class DouyinPageSearchTests(unittest.IsolatedAsyncioTestCase):
    async def test_verification_wait_is_passive_and_page_response_resumes_collection(self):
        verify_payload = {
            "data": [],
            "search_nil_info": {"search_nil_type": "verify_check"},
        }
        success_payload = {"data": [{"aweme_info": {
            "aweme_id": "page-result-1",
            "desc": "页面自然响应",
            "video": {"play_addr": {"url_list": ["https://media.test/1.mp4"]}},
        }}]}

        class Response:
            status = 200
            url = "https://www.douyin.com/aweme/v1/web/general/search/single/"

            def __init__(self, payload):
                self.payload = payload

            async def json(self):
                return self.payload

        class Page:
            def __init__(self):
                self.url = "about:blank"
                self.handler = None
                self.wait_calls = 0
                self.evaluate_calls = 0
                self.mouse = SimpleNamespace(wheel=AsyncMock())

            def on(self, event, handler):
                if event == "response":
                    self.handler = handler

            async def goto(self, url, **_kwargs):
                self.url = url
                await self.handler(Response(verify_payload))

            async def wait_for_load_state(self, *_args, **_kwargs):
                return None

            async def wait_for_timeout(self, _timeout):
                self.wait_calls += 1
                # 第一次是页面 settle；第二次是验证码被动等待。模拟用户完成验证后，
                # 页面自己产生正常搜索响应。
                if self.wait_calls == 2:
                    await self.handler(Response(success_payload))

            async def bring_to_front(self):
                return None

            async def evaluate(self, *_args, **_kwargs):
                self.evaluate_calls += 1

            async def close(self):
                return None

        page = Page()
        context = SimpleNamespace(pages=[page])

        rows, error = await fetch_douyin_search(
            SimpleNamespace(), SimpleNamespace(), "露营",
            max_results=1, captcha_wait_seconds=3, context=context,
        )

        self.assertEqual([row["aweme_id"] for row in rows], ["page-result-1"])
        self.assertEqual(error, "")
        self.assertEqual(page.evaluate_calls, 0)
        self.assertEqual(page.mouse.wheel.await_count, 0)


class KeywordCollectionPipelineTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "collection.db"
        db.init_db(str(self.db_path))
        with db.get_session() as session:
            account = DouyinAccount(
                platform="douyin",
                nickname="collector",
                status="active",
                storage_state=json.dumps({
                    "cookies": [{
                        "name": "sessionid",
                        "value": "fixture-session",
                        "domain": ".douyin.com",
                        "path": "/",
                    }],
                    "origins": [],
                }),
            )
            session.add(account)
            session.commit()
            session.refresh(account)
            self.account = SimpleNamespace(
                id=account.id,
                platform=account.platform,
                storage_state=account.storage_state,
                proxy=account.proxy,
            )
            job = KeywordCollectionJob(
                platform="douyin",
                account_id=account.id,
                keywords=json.dumps(["露营"], ensure_ascii=False),
                max_contents_per_keyword=2,
                max_comments_per_content=2,
            )
            session.add(job)
            session.commit()
            session.refresh(job)
            self.job_id = job.id

    async def asyncTearDown(self):
        if db._engine is not None:
            db._engine.dispose()
        self.temp_dir.cleanup()

    async def test_pipeline_persists_content_comments_and_deduplicates_retry(self):
        raw_aweme = {
            "aweme_id": "aweme-1",
            "desc": "周末露营清单",
            "create_time": 1_700_000_000,
            "author": {"nickname": "示例作者", "sec_uid": "author-1"},
            "statistics": {"digg_count": 88, "comment_count": 12},
            "video": {
                "play_addr": {"url_list": ["https://media.test/aweme-1.mp4"]},
                "cover": {"url_list": ["https://media.test/cover.jpg"]},
            },
        }
        comments = [
            {"comment_id": "c-1", "text": "想要链接", "user_nickname": "用户甲",
             "like_count": 3, "create_time": 1_700_000_001, "reply_to": ""},
            {"comment_id": "c-2", "text": "已收藏", "user_nickname": "用户乙",
             "like_count": 1, "create_time": 1_700_000_002, "reply_to": ""},
        ]
        headed_calls = []

        @asynccontextmanager
        async def temporary_headed_context(identity):
            headed_calls.append(identity)
            yield SimpleNamespace()

        browser = SimpleNamespace(
            identity_for=lambda _account: SimpleNamespace(ua="fixture-agent"),
            temporary_headed_context=temporary_headed_context,
        )
        collector = KeywordCollector(Config(), browser, SimpleNamespace())
        collector._discover_douyin = AsyncMock(return_value=([raw_aweme], ""))
        collector._douyin_comments = AsyncMock(return_value=(comments, ""))

        first = await collector.run(self.job_id, self.account)
        second = await collector.run(self.job_id, self.account)

        self.assertEqual(first["contents"], 1)
        self.assertEqual(first["comments"], 2)
        self.assertEqual(second["contents"], 1)
        self.assertEqual(second["comments"], 2)
        with db.get_session() as session:
            contents = session.exec(select(KeywordCollectionContent)).all()
            saved_comments = session.exec(select(KeywordCollectionComment)).all()
            job = session.get(KeywordCollectionJob, self.job_id)
        self.assertEqual(len(contents), 1)
        self.assertEqual(len(saved_comments), 2)
        self.assertEqual(contents[0].author_id, "author-1")
        self.assertEqual(contents[0].collected_comment_count, 2)
        self.assertEqual(job.content_count, 1)
        self.assertEqual(job.comment_count, 2)
        self.assertEqual(len(headed_calls), 2)

    async def test_captcha_stops_remaining_keywords_in_same_job(self):
        with db.get_session() as session:
            job = session.get(KeywordCollectionJob, self.job_id)
            job.keywords = json.dumps(["露营", "穿搭"], ensure_ascii=False)
            session.add(job)
            session.commit()

        @asynccontextmanager
        async def temporary_headed_context(_identity):
            yield SimpleNamespace()

        browser = SimpleNamespace(
            identity_for=lambda _account: SimpleNamespace(ua="fixture-agent"),
            temporary_headed_context=temporary_headed_context,
        )
        cfg = Config()
        cfg.engine.douyin_keyword_gap_seconds = 0
        collector = KeywordCollector(cfg, browser, SimpleNamespace())
        collector._discover_douyin = AsyncMock(return_value=(
            [], "抖音要求完成滑块验证；本次任务已停止后续请求",
        ))

        result = await collector.run(self.job_id, self.account)

        self.assertEqual(collector._discover_douyin.await_count, 1)
        self.assertEqual(result["errors"], 1)
        with db.get_session() as session:
            job = session.get(KeywordCollectionJob, self.job_id)
        self.assertIn("滑块验证", job.error)


class KeywordCollectionEditTests(unittest.TestCase):
    def setUp(self):
        self.previous_engine = db._engine
        self.temp_dir = tempfile.TemporaryDirectory()
        db.init_db(str(Path(self.temp_dir.name) / "collection-edit.db"))
        with db.get_session() as session:
            account = DouyinAccount(
                platform="douyin", nickname="编辑账号", status="active",
                storage_state='{"cookies":[{"name":"sessionid","value":"fixture"}]}',
            )
            session.add(account); session.commit(); session.refresh(account)
            job = KeywordCollectionJob(
                platform="douyin", account_id=account.id,
                keywords='["旧关键词"]', status="done", current_step="已完成",
            )
            session.add(job); session.commit(); session.refresh(job)
            content = KeywordCollectionContent(
                job_id=job.id, platform="douyin", keyword="旧关键词",
                aweme_id="keep-result",
            )
            session.add(content); session.commit()
            self.account_id, self.job_id = account.id, job.id

    def tearDown(self):
        if db._engine is not None:
            db._engine.dispose()
        db._engine = self.previous_engine
        self.temp_dir.cleanup()

    def _body(self, **changes):
        values = dict(
            platform="douyin", account_id=self.account_id,
            keywords=["新关键词", "第二个词"],
            max_contents_per_keyword=12,
            max_pages_per_keyword=24, stagnant_pages=4,
            search_sort="latest", publish_time="week", content_type="video",
            min_likes=100, min_comments=5,
            max_comments_per_content=30,
            include_replies=True, download_media=True,
            video_quality="1080", download_dir="",
        )
        values.update(changes)
        return KeywordCollectionIn(**values)

    def test_finished_task_can_be_edited_without_deleting_results(self):
        result = asyncio.run(update_keyword_collection(self.job_id, self._body()))

        self.assertEqual(result["keywords"], ["新关键词", "第二个词"])
        self.assertEqual(result["max_contents_per_keyword"], 12)
        self.assertEqual(result["max_pages_per_keyword"], 24)
        self.assertEqual(result["stagnant_pages"], 4)
        self.assertEqual(result["search_sort"], "latest")
        self.assertEqual(result["publish_time"], "week")
        self.assertEqual(result["content_type"], "video")
        self.assertEqual(result["min_likes"], 100)
        self.assertEqual(result["min_comments"], 5)
        self.assertEqual(result["max_comments_per_content"], 30)
        self.assertTrue(result["include_replies"])
        self.assertTrue(result["download_media"])
        self.assertEqual(result["video_quality"], "1080")
        self.assertEqual(result["status"], "done")
        self.assertEqual(result["current_step"], "配置已更新，可点击续跑")
        with db.get_session() as session:
            kept = session.exec(select(KeywordCollectionContent)).all()
        self.assertEqual([row.aweme_id for row in kept], ["keep-result"])

    def test_running_or_pending_task_must_be_canceled_before_edit(self):
        with db.get_session() as session:
            job = session.get(KeywordCollectionJob, self.job_id)
            job.status = "running"
            session.add(job); session.commit()

        with self.assertRaises(HTTPException) as caught:
            asyncio.run(update_keyword_collection(self.job_id, self._body()))
        self.assertEqual(caught.exception.status_code, 409)

    def test_collection_depth_and_filter_values_are_validated(self):
        invalid_values = (
            {"max_pages_per_keyword": 0},
            {"stagnant_pages": 9},
            {"search_sort": "unknown"},
            {"publish_time": "month"},
            {"content_type": "live"},
            {"min_likes": -1},
        )
        for values in invalid_values:
            with self.subTest(values=values), self.assertRaises(HTTPException) as caught:
                asyncio.run(update_keyword_collection(self.job_id, self._body(**values)))
            self.assertEqual(caught.exception.status_code, 400)

    def test_xhs_collection_creation_is_deferred(self):
        with self.assertRaises(HTTPException) as caught:
            asyncio.run(create_keyword_collection(_req(), self._body(platform="xhs")))
        self.assertEqual(caught.exception.status_code, 400)
        self.assertIn("仅支持抖音", caught.exception.detail)


class KeywordCollectionMediaPreviewTests(unittest.TestCase):
    def setUp(self):
        self.previous_engine = db._engine
        self.temp_dir = tempfile.TemporaryDirectory()
        db.init_db(str(Path(self.temp_dir.name) / "collection-preview.db"))
        with db.get_session() as session:
            job = KeywordCollectionJob(
                platform="douyin", account_id=1, keywords='["预览"]', status="done",
            )
            session.add(job); session.commit(); session.refresh(job)
            self.job_id = job.id

    def tearDown(self):
        if db._engine is not None:
            db._engine.dispose()
        db._engine = self.previous_engine
        self.temp_dir.cleanup()

    def _add_content(self, **values):
        defaults = dict(
            job_id=self.job_id, platform="douyin", keyword="预览",
            aweme_id="preview-1", desc="可预览视频", media_type="video",
            media_json=json.dumps([{
                "kind": "video", "ext": "mp4",
                "url": "https://media.test/preview-1.mp4",
            }]),
            download_status="done",
        )
        defaults.update(values)
        with db.get_session() as session:
            content = KeywordCollectionContent(**defaults)
            session.add(content); session.commit(); session.refresh(content)
            return content.id

    def test_downloaded_video_is_exposed_to_result_list_and_preview(self):
        media = Path(self.temp_dir.name) / "preview-1_title.mp4"
        media.write_bytes(b"fixture-video")
        content_id = self._add_content(local_path=str(media))

        with db.get_session() as session:
            item = _collection_content_dict(session.get(KeywordCollectionContent, content_id))
        preview = asyncio.run(keyword_collection_content_media(self.job_id, content_id))
        response = asyncio.run(keyword_collection_local_media(self.job_id, content_id, 0))

        self.assertTrue(item["local_exists"])
        self.assertTrue(item["preview_available"])
        self.assertEqual(item["media_count"], 1)
        self.assertEqual(item["file_size"], len(b"fixture-video"))
        self.assertEqual(
            preview["local_url"],
            f"/api/collections/{self.job_id}/contents/{content_id}/local-media/0",
        )
        self.assertEqual(preview["medias"][0]["kind"], "video")
        self.assertEqual(Path(response.path), media.resolve())

    def test_gallery_preview_only_includes_files_for_current_aweme(self):
        folder = Path(self.temp_dir.name) / "author"
        folder.mkdir()
        (folder / "gallery-1_title_1.jpg").write_bytes(b"second")
        (folder / "gallery-1_title_0.jpg").write_bytes(b"first")
        (folder / "another-work_title_0.jpg").write_bytes(b"other")
        content_id = self._add_content(
            aweme_id="gallery-1", media_type="images", local_path=str(folder),
            media_json="[]",
        )

        preview = asyncio.run(keyword_collection_content_media(self.job_id, content_id))

        self.assertEqual(len(preview["medias"]), 2)
        self.assertTrue(all(item["kind"] == "image" for item in preview["medias"]))
        self.assertTrue(preview["medias"][0]["url"].endswith("/local-media/0"))
        self.assertEqual(preview["local_url"], "")

    def test_preview_rejects_content_from_another_job(self):
        content_id = self._add_content()
        with self.assertRaises(HTTPException) as caught:
            asyncio.run(keyword_collection_content_media(self.job_id + 99, content_id))
        self.assertEqual(caught.exception.status_code, 404)


class KeywordCollectionReportTests(unittest.TestCase):
    def test_report_has_overview_content_and_comment_sheets(self):
        job = KeywordCollectionJob(
            id=7,
            platform="xhs",
            account_id=1,
            keywords='["露营"]',
            status="done",
        )
        content = KeywordCollectionContent(
            id=8,
            job_id=7,
            platform="xhs",
            keyword="露营",
            aweme_id="note-1",
            desc="=FORMULA",
            collected_comment_count=1,
        )
        comment = KeywordCollectionComment(
            id=9,
            job_id=7,
            content_id=8,
            platform="xhs",
            aweme_id="note-1",
            comment_id="comment-1",
            text="+comment",
        )

        workbook = load_workbook(BytesIO(
            build_keyword_collection_report(job, [content], [comment])
        ))

        self.assertEqual(workbook.sheetnames, ["概览", "作品", "评论"])
        self.assertEqual(workbook["作品"]["E2"].value, "'=FORMULA")
        self.assertEqual(workbook["评论"]["F2"].value, "'+comment")


if __name__ == "__main__":
    unittest.main()
