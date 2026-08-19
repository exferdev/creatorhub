"""Excel report generation for monitored platform data."""
from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from io import BytesIO
import json
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins


REPORT_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_HEADER_FILL = PatternFill("solid", fgColor="17324D")
_TITLE_FILL = PatternFill("solid", fgColor="0F766E")
_SUBTITLE_FILL = PatternFill("solid", fgColor="E6FFFB")
_META_LABEL_FILL = PatternFill("solid", fgColor="EAF2F8")
_META_VALUE_FILL = PatternFill("solid", fgColor="F8FAFC")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F7FAFC")
_KPI_FILL = PatternFill("solid", fgColor="F0FDFA")
_LINK_FILL = PatternFill("solid", fgColor="EFF6FF")
_SUCCESS_FILL = PatternFill("solid", fgColor="ECFDF5")
_INFO_FILL = PatternFill("solid", fgColor="EFF6FF")
_WARN_FILL = PatternFill("solid", fgColor="FFFBEB")
_DANGER_FILL = PatternFill("solid", fgColor="FEF2F2")
_MUTED_FILL = PatternFill("solid", fgColor="F3F4F6")
_HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
_TITLE_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
_BODY_FONT = Font(name="Microsoft YaHei", size=10)
_LABEL_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="334155")
_LINK_FONT = Font(name="Microsoft YaHei", size=10, color="2563EB", underline="single")
_SUCCESS_FONT = Font(name="Microsoft YaHei", size=10, color="047857", bold=True)
_INFO_FONT = Font(name="Microsoft YaHei", size=10, color="1D4ED8", bold=True)
_WARN_FONT = Font(name="Microsoft YaHei", size=10, color="B45309", bold=True)
_DANGER_FONT = Font(name="Microsoft YaHei", size=10, color="B91C1C", bold=True)
_MUTED_FONT = Font(name="Microsoft YaHei", size=10, color="64748B")
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="D9E2EC"),
)
_WRAP = Alignment(vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_NUMBER = Alignment(horizontal="right", vertical="top")
_PAGE_MARGINS = PageMargins(left=0.25, right=0.25, top=0.45, bottom=0.45, header=0.2, footer=0.2)


def _value(obj: Any, name: str, default: Any = "") -> Any:
    if isinstance(obj, dict):
        return obj.get(name, default)
    return getattr(obj, name, default)


def _text(value: Any) -> str:
    """Turn user-controlled text into a harmless Excel string.

    A leading formula character in a description or comment must not be
    interpreted as an Excel formula when the exported file is opened.
    """
    if value is None:
        return ""
    result = str(value)
    if result[:1] in ("=", "+", "-", "@"):
        return "'" + result
    return result


def _excel_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone().replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return None


def _timestamp(value: Any) -> datetime | None:
    try:
        seconds = int(value or 0)
    except (TypeError, ValueError):
        return None
    if seconds <= 0:
        return None
    try:
        return datetime.fromtimestamp(seconds)
    except (OverflowError, OSError, ValueError):
        return None


def _tags(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        return ", ".join(_text(item) for item in value if str(item).strip())
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return ", ".join(_text(item) for item in parsed if str(item).strip())
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    return _text(value)


def _local_path_text(value: Any) -> str:
    """Expose a local file reference as an absolute path in the workbook."""
    raw = _text(value).strip()
    if not raw or raw.startswith(("http://", "https://", "file://")):
        return raw
    try:
        return str(Path(raw).expanduser().resolve())
    except (OSError, RuntimeError, ValueError):
        return raw


def _local_hyperlink_formula(value: str) -> str:
    """Build a WPS-compatible link to a Windows file or folder.

    An OOXML hyperlink relationship stores non-ASCII Windows paths as an
    encoded URI.  Some WPS desktop versions do not decode that relationship
    consistently.  HYPERLINK keeps the absolute path in the worksheet
    formula, which WPS supports for both files and folders.
    """
    escaped = value.replace('"', '""')
    return f'=HYPERLINK("{escaped}","{escaped}")'


def _configure_workbook(workbook: Workbook) -> None:
    """Make formula-based links recalculate when WPS opens the workbook."""
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _internal_hyperlink(cell: Any, sheet_title: str, target: str = "A1") -> None:
    """Create an internal hyperlink using WPS/Excel's location field."""
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{_safe_sheet_title(sheet_title)}'!{target}",
        display=str(cell.value or ""),
    )


def _label(item: Any, *, fallback: str = "") -> str:
    return _text(
        _value(item, "alias")
        or _value(item, "nickname")
        or _value(item, "title")
        or _value(item, "keyword")
        or _value(item, "sec_uid")
        or _value(item, "aweme_id")
        or fallback
    )


def _target_map(items: Iterable[Any]) -> dict[Any, str]:
    return {
        _value(item, "id"): _label(item)
        for item in items
        if _value(item, "id") is not None
    }


def _watch_map(items: Iterable[Any]) -> dict[Any, str]:
    return {
        _value(item, "id"): _label(item)
        for item in items
        if _value(item, "id") is not None
    }


def _safe_sheet_title(value: str) -> str:
    # Excel forbids these characters in worksheet titles.
    return "".join(ch for ch in value if ch not in "[]:*?/\\")[:31] or "数据"


def _status_style(cell: Any, header: str, value: Any) -> None:
    """Apply restrained semantic colors without making color the only signal."""
    text = str(value or "").strip().casefold()
    header = str(header or "").strip()
    if header in {"下载状态", "状态"}:
        styles = {
            "done": (_SUCCESS_FILL, _SUCCESS_FONT),
            "成功": (_SUCCESS_FILL, _SUCCESS_FONT),
            "downloading": (_INFO_FILL, _INFO_FONT),
            "处理中": (_INFO_FILL, _INFO_FONT),
            "pending": (_WARN_FILL, _WARN_FONT),
            "等待中": (_WARN_FILL, _WARN_FONT),
            "failed": (_DANGER_FILL, _DANGER_FONT),
            "失败": (_DANGER_FILL, _DANGER_FONT),
            "skipped": (_MUTED_FILL, _MUTED_FONT),
            "仅记录": (_MUTED_FILL, _MUTED_FONT),
        }
        if text in styles:
            cell.fill, cell.font = styles[text]
    elif header == "启用":
        cell.fill, cell.font = (
            (_SUCCESS_FILL, _SUCCESS_FONT)
            if text in {"是", "true", "1"}
            else (_MUTED_FILL, _MUTED_FONT)
        )
    elif header == "已屏蔽":
        cell.fill, cell.font = (
            (_WARN_FILL, _WARN_FONT)
            if text in {"是", "true", "1"}
            else (_SUCCESS_FILL, _SUCCESS_FONT)
        )
    elif header in {"错误", "最近错误"} and text:
        cell.fill, cell.font = _DANGER_FILL, _DANGER_FONT


def _configure_sheet(
    worksheet: Any,
    *,
    tab_color: str | None,
    freeze_panes: str,
    landscape: bool = True,
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = freeze_panes
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape" if landscape else "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = _PAGE_MARGINS
    worksheet.sheet_properties.tabColor = tab_color or "17324D"


def _write_table(
    workbook: Workbook,
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    widths: Sequence[int] | None = None,
    date_columns: set[int] | None = None,
    number_columns: set[int] | None = None,
    number_formats: Mapping[int, str] | None = None,
    link_columns: set[int] | None = None,
    tab_color: str | None = None,
    freeze_panes: str = "A2",
    row_height: float = 28,
) -> None:
    worksheet = workbook.create_sheet(_safe_sheet_title(title))
    _configure_sheet(
        worksheet,
        tab_color=tab_color,
        freeze_panes=freeze_panes,
    )
    worksheet.sheet_format.defaultRowHeight = row_height
    number_formats = number_formats or {}
    link_columns = link_columns or set()

    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    worksheet.row_dimensions[1].height = 30

    materialized = list(rows)
    date_columns = date_columns or set()
    number_columns = number_columns or set()
    for row_number, row in enumerate(materialized, 2):
        row_text_lengths = [len(str(value)) for value in row if value not in (None, "")]
        worksheet.row_dimensions[row_number].height = min(
            60,
            max(row_height, 42 if any(length > 90 for length in row_text_lengths) else row_height),
        )
        for col, value in enumerate(row, 1):
            if isinstance(value, str):
                value = _text(value)
            cell = worksheet.cell(row=row_number, column=col, value=value)
            cell.font = _BODY_FONT
            cell.alignment = _NUMBER if col in number_columns else _WRAP
            cell.border = _THIN_BORDER
            if row_number % 2 == 0:
                cell.fill = _ALT_ROW_FILL
            if col in date_columns and value:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
                cell.alignment = _CENTER
            if col in number_formats:
                cell.number_format = number_formats[col]
            if col in link_columns and isinstance(value, str) and value:
                link_ready = False
                if value.startswith(("http://", "https://", "file://")):
                    cell.hyperlink = value
                    link_ready = True
                elif headers[col - 1] in {"保存目录", "本地文件", "文件路径"}:
                    try:
                        target = str(Path(value).expanduser().resolve())
                        cell.value = _local_hyperlink_formula(target)
                        link_ready = True
                    except (OSError, RuntimeError, ValueError):
                        pass
                if link_ready:
                    cell.font = _LINK_FONT
                    cell.fill = _LINK_FILL
            _status_style(cell, headers[col - 1] if col <= len(headers) else "", value)

    if widths:
        for col, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width
    else:
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 16

    last_row = max(1, worksheet.max_row)
    last_col = max(1, worksheet.max_column)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    worksheet.print_area = worksheet.auto_filter.ref
    worksheet.oddFooter.center.text = "&A  ·  第 &P / &N 页"
    if materialized:
        table_name = "Table_" + "".join(ch for ch in title if ch.isalnum())
        table_name = (table_name or "Table_Data")[:240]
        # Workbook table names must be unique even if a future caller uses
        # two titles that normalize to the same identifier.
        table_name += "_" + str(len(workbook.worksheets) + 1)
        table = Table(displayName=table_name, ref=worksheet.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def _write_module_summary(
    workbook: Workbook,
    *,
    report_title: str,
    module_name: str,
    detail_sheet: str,
    record_count: int,
    filters: Sequence[tuple[str, Any]],
    generated_at: datetime,
) -> None:
    worksheet = workbook.active
    worksheet.title = "概览"
    _configure_sheet(
        worksheet,
        tab_color="0F766E",
        freeze_panes="A3",
        landscape=False,
    )
    worksheet.sheet_format.defaultRowHeight = 22
    worksheet.merge_cells("A1:F1")
    title = worksheet["A1"]
    title.value = report_title
    title.fill = _TITLE_FILL
    title.font = _TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30
    for column in range(1, 7):
        worksheet.cell(row=1, column=column).fill = _TITLE_FILL

    worksheet.merge_cells("A2:F2")
    worksheet["A2"] = "数据导出概览 · 明细工作表已启用筛选、冻结表头与交替行"
    worksheet["A2"].font = Font(name="Microsoft YaHei", size=10, color="0F766E")
    worksheet["A2"].fill = _SUBTITLE_FILL
    worksheet["A2"].alignment = _WRAP

    metadata = [
        ("模块", module_name),
        ("记录数", record_count),
        ("生成时间", generated_at),
    ]
    for row, (key, value) in enumerate(metadata, 3):
        label = worksheet.cell(row=row, column=1, value=key)
        label.font = _LABEL_FONT
        label.fill = _META_LABEL_FILL
        label.alignment = _CENTER
        label.border = _THIN_BORDER
        cell = worksheet.cell(row=row, column=2, value=value)
        cell.font = _BODY_FONT
        cell.alignment = _WRAP
        cell.fill = _KPI_FILL if key == "记录数" else _META_VALUE_FILL
        cell.border = _THIN_BORDER
        if key == "记录数":
            cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="0F766E")
            cell.number_format = "#,##0"
        if isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    worksheet["A7"] = "导出条件"
    worksheet["A7"].fill = _SUBTITLE_FILL
    worksheet["A7"].font = Font(name="Microsoft YaHei", bold=True, color="0F766E")
    worksheet["A7"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.merge_cells("A7:F7")
    for column in range(1, 7):
        worksheet.cell(row=7, column=column).fill = _SUBTITLE_FILL
    if not filters:
        filters = (("筛选", "全部数据"),)
    filter_rows = (len(filters) + 1) // 2
    for index, (key, value) in enumerate(filters):
        row = 8 + index // 2
        column = 1 + (index % 2) * 3
        label = worksheet.cell(row=row, column=column, value=key)
        label.font = _LABEL_FONT
        label.fill = _META_LABEL_FILL
        label.alignment = _CENTER
        label.border = _THIN_BORDER
        worksheet.merge_cells(
            start_row=row,
            start_column=column + 1,
            end_row=row,
            end_column=column + 2,
        )
        value_cell = worksheet.cell(row=row, column=column + 1, value=_text(value) or "全部")
        value_cell.font = _BODY_FONT
        value_cell.alignment = _WRAP
        value_cell.fill = _META_VALUE_FILL
        value_cell.border = _THIN_BORDER

    detail_row = 8 + max(filter_rows, 1) + 2
    worksheet.cell(row=detail_row, column=1, value="明细工作表").font = _LABEL_FONT
    worksheet.cell(row=detail_row, column=1).fill = _SUBTITLE_FILL
    worksheet.cell(row=detail_row, column=1).border = _THIN_BORDER
    worksheet.merge_cells(start_row=detail_row, start_column=2, end_row=detail_row, end_column=6)
    detail = worksheet.cell(row=detail_row, column=2, value=f"查看《{detail_sheet}》")
    detail.font = _LINK_FONT
    detail.fill = _LINK_FILL
    detail.border = _THIN_BORDER
    _internal_hyperlink(detail, detail_sheet)

    for column, width in {"A": 18, "B": 26, "C": 16, "D": 18, "E": 26, "F": 16}.items():
        worksheet.column_dimensions[column].width = width
    worksheet.print_area = f"A1:F{detail_row}"
    worksheet.oddFooter.center.text = "&A  ·  第 &P / &N 页"


def build_module_report(
    *,
    report_title: str,
    module_name: str,
    sheet_title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    filters: Sequence[tuple[str, Any]] = (),
    widths: Sequence[int] | None = None,
    date_columns: set[int] | None = None,
    number_columns: set[int] | None = None,
    number_formats: Mapping[int, str] | None = None,
    link_columns: set[int] | None = None,
    tab_color: str | None = None,
    freeze_panes: str = "A2",
    row_height: float = 28,
    generated_at: datetime | None = None,
) -> bytes:
    """Build a workbook containing exactly one filtered module dataset."""
    materialized = list(rows)
    workbook = Workbook()
    _configure_workbook(workbook)
    generated_at = generated_at or datetime.now()
    workbook.properties.creator = "CreatorHub"
    workbook.properties.title = report_title
    workbook.properties.subject = f"{module_name} Excel 导出"
    _write_module_summary(
        workbook,
        report_title=report_title,
        module_name=module_name,
        detail_sheet=sheet_title,
        record_count=len(materialized),
        filters=filters,
        generated_at=generated_at,
    )
    _write_table(
        workbook,
        sheet_title,
        headers,
        materialized,
        widths=widths,
        date_columns=date_columns,
        number_columns=number_columns,
        number_formats=number_formats,
        link_columns=link_columns,
        tab_color=tab_color,
        freeze_panes=freeze_panes,
        row_height=row_height,
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


_CONTENT_HEADERS = (
    "ID", "监控目标", "平台", "作品ID", "描述", "媒体类型", "发布时间",
    "采集时间", "点赞数", "评论数", "时长(秒)", "画质", "下载状态", "本地文件",
    "封面链接", "错误",
)
_COMMENT_HEADERS = (
    "ID", "评论监控", "平台", "作品ID", "评论ID", "评论内容", "用户昵称",
    "点赞数", "评论时间", "采集时间", "类型", "上级评论ID",
)
_DANMAKU_HEADERS = (
    "ID", "弹幕监控", "平台", "作品ID", "弹幕ID", "弹幕内容", "用户ID",
    "用户昵称", "视频内时间(秒)", "发送时间", "采集时间", "点赞数", "来源", "已屏蔽",
)
_SHARE_HISTORY_HEADERS = (
    "ID", "平台", "作品ID", "作品描述", "作者", "媒体类型", "媒体数量",
    "发布时间", "下载时间", "点赞数", "评论数", "时长(秒)", "画质", "下载状态",
    "保存目录", "本地文件", "来源链接", "错误",
)
_TARGET_HEADERS = (
    "ID", "平台", "类型", "目标名称", "别名", "分组", "标签", "启用",
    "检查间隔(秒)", "最后扫描", "作品数", "最近错误",
)
_WATCH_HEADERS = (
    "ID", "平台", "对象类型", "监控对象", "别名", "分组", "标签", "评论来源",
    "启用", "评论数", "检查间隔(秒)", "最后扫描", "最近错误",
)
_DANMAKU_WATCH_HEADERS = (
    "ID", "平台", "对象类型", "监控对象", "别名", "分组", "标签", "弹幕来源",
    "启用", "弹幕数", "检查间隔(秒)", "最后扫描", "最近错误",
)


def _content_row(row: Any, target_names: dict[Any, str]) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        target_names.get(_value(row, "target_id"), f"目标#{_value(row, 'target_id')}"),
        _text(_value(row, "platform")),
        _text(_value(row, "aweme_id")),
        _text(_value(row, "desc")),
        _text(_value(row, "media_type")),
        _timestamp(_value(row, "create_time")),
        _excel_datetime(_value(row, "created_at")),
        _value(row, "like_count", 0),
        _value(row, "comment_count", 0),
        _value(row, "duration", 0),
        _text(_value(row, "quality")),
        _text(_value(row, "download_status")),
        _local_path_text(_value(row, "local_path")),
        _text(_value(row, "cover_url")),
        _text(_value(row, "error")),
    )


def _comment_row(row: Any, watch_names: dict[Any, str]) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        watch_names.get(_value(row, "watch_id"), f"监控#{_value(row, 'watch_id')}"),
        _text(_value(row, "platform")),
        _text(_value(row, "aweme_id")),
        _text(_value(row, "comment_id")),
        _text(_value(row, "text")),
        _text(_value(row, "user_nickname")),
        _value(row, "like_count", 0),
        _timestamp(_value(row, "create_time")),
        _excel_datetime(_value(row, "created_at")),
        "回复" if _value(row, "reply_to") else "一级评论",
        _text(_value(row, "reply_to")),
    )


def _danmaku_row(row: Any, watch_names: dict[Any, str]) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        watch_names.get(_value(row, "watch_id"), f"监控#{_value(row, 'watch_id')}"),
        _text(_value(row, "platform")),
        _text(_value(row, "aweme_id")),
        _text(_value(row, "danmaku_id")),
        _text(_value(row, "text")),
        _text(_value(row, "user_id")),
        _text(_value(row, "user_nickname")),
        round(int(_value(row, "video_time_ms", 0) or 0) / 1000, 3),
        _timestamp(_value(row, "create_time")),
        _excel_datetime(_value(row, "created_at")),
        _value(row, "like_count", 0),
        _text(_value(row, "source")),
        "是" if _value(row, "is_blocked", False) else "否",
    )


def _target_row(row: Any, content_count: int) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        _text(_value(row, "platform")),
        _text(_value(row, "target_kind")),
        _label(row),
        _text(_value(row, "alias")),
        _text(_value(row, "group_name")),
        _tags(_value(row, "tags")),
        "是" if _value(row, "enabled", True) else "否",
        _value(row, "interval_seconds", 0),
        _excel_datetime(_value(row, "last_scan_at")),
        content_count,
        _text(_value(row, "last_error")),
    )


def _watch_row(row: Any) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        _text(_value(row, "platform")),
        _text(_value(row, "kind")),
        _label(row),
        _text(_value(row, "alias")),
        _text(_value(row, "group_name")),
        _tags(_value(row, "tags")),
        _text(_value(row, "mode")),
        "是" if _value(row, "enabled", True) else "否",
        _value(row, "comment_count", 0),
        _value(row, "interval_seconds", 0),
        _excel_datetime(_value(row, "last_scan_at")),
        _text(_value(row, "last_error")),
    )


def _danmaku_watch_row(row: Any) -> tuple[Any, ...]:
    return (
        _value(row, "id"),
        _text(_value(row, "platform")),
        _text(_value(row, "kind")),
        _label(row),
        _text(_value(row, "alias")),
        _text(_value(row, "group_name")),
        _tags(_value(row, "tags")),
        _text(_value(row, "mode")),
        "是" if _value(row, "enabled", True) else "否",
        _value(row, "danmaku_count", 0),
        _value(row, "interval_seconds", 0),
        _excel_datetime(_value(row, "last_scan_at")),
        _text(_value(row, "last_error")),
    )


def _share_history_row(row: Any) -> tuple[Any, ...]:
    metadata = _value(row, "metadata", {})
    if not isinstance(metadata, dict):
        metadata = {}
    files = _value(row, "files", [])
    if not isinstance(files, list):
        files = []
    media_files = [item for item in files if isinstance(item, dict) and item.get("role") == "media"]
    media_count = _value(row, "media_count", 0) or metadata.get("media_count") or len(media_files)
    try:
        media_count = int(media_count or 0)
    except (TypeError, ValueError):
        media_count = 0
    local_paths = [
        _local_path_text(item.get("path"))
        for item in media_files
        if isinstance(item, dict) and item.get("path")
    ]
    local_path = _value(row, "local_path", "") or "\n".join(local_paths)
    return (
        _value(row, "id"),
        _text(_value(row, "platform")),
        _text(_value(row, "item_id") or _value(row, "aweme_id")),
        _text(_value(row, "title") or _value(row, "desc") or metadata.get("title")),
        _text(_value(row, "author") or metadata.get("uploader") or metadata.get("channel")),
        _text(_value(row, "media_type") or metadata.get("media_type")),
        media_count,
        _timestamp(_value(row, "create_time") or metadata.get("timestamp")),
        _excel_datetime(_value(row, "created_at")),
        _value(row, "like_count", 0) or metadata.get("like_count", 0) or 0,
        _value(row, "comment_count", 0) or metadata.get("comment_count", 0) or 0,
        _value(row, "duration", 0) or metadata.get("duration", 0) or 0,
        _text(_value(row, "quality") or metadata.get("format") or metadata.get("format_id")),
        _text(_value(row, "status") or _value(row, "download_status")),
        _local_path_text(_value(row, "output_dir")),
        _local_path_text(local_path),
        _text(_value(row, "source_url")),
        _text(_value(row, "error")),
    )


def build_contents_report(
    contents: Sequence[Any],
    targets: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    target_names = _target_map(targets)
    return build_module_report(
        report_title="CreatorHub 作品数据报告",
        module_name="作品数据",
        sheet_title="作品数据",
        headers=_CONTENT_HEADERS,
        rows=(_content_row(row, target_names) for row in contents),
        filters=filters,
        widths=(10, 28, 12, 24, 50, 12, 20, 20, 12, 12, 12, 12, 14, 40, 40, 36),
        date_columns={7, 8},
        number_columns={1, 9, 10, 11},
        number_formats={1: "0", 9: "#,##0", 10: "#,##0", 11: "0.0"},
        link_columns={14, 15},
        tab_color="2563EB",
        freeze_panes="E2",
        row_height=34,
        generated_at=generated_at,
    )


def build_comments_report(
    comments: Sequence[Any],
    watches: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    watch_names = _watch_map(watches)
    return build_module_report(
        report_title="CreatorHub 评论数据报告",
        module_name="评论数据",
        sheet_title="评论数据",
        headers=_COMMENT_HEADERS,
        rows=(_comment_row(row, watch_names) for row in comments),
        filters=filters,
        widths=(10, 28, 12, 24, 24, 56, 22, 12, 20, 20, 14, 24),
        date_columns={9, 10},
        number_columns={1, 8},
        number_formats={1: "0", 8: "#,##0"},
        tab_color="14B8A6",
        freeze_panes="F2",
        row_height=36,
        generated_at=generated_at,
    )


def build_danmaku_report(
    danmaku: Sequence[Any],
    watches: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    watch_names = _watch_map(watches)
    return build_module_report(
        report_title="CreatorHub 弹幕数据报告",
        module_name="弹幕数据",
        sheet_title="弹幕数据",
        headers=_DANMAKU_HEADERS,
        rows=(_danmaku_row(row, watch_names) for row in danmaku),
        filters=filters,
        widths=(10, 28, 12, 24, 24, 56, 24, 22, 18, 20, 20, 12, 12, 12),
        date_columns={10, 11},
        number_columns={1, 9, 12},
        number_formats={1: "0", 9: "0.000", 12: "#,##0"},
        tab_color="8B5CF6",
        freeze_panes="F2",
        row_height=36,
        generated_at=generated_at,
    )


def build_targets_report(
    targets: Sequence[Any],
    contents: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    counts = Counter(_value(row, "target_id") for row in contents)
    return build_module_report(
        report_title="CreatorHub 监控目标报告",
        module_name="监控目标",
        sheet_title="监控目标",
        headers=_TARGET_HEADERS,
        rows=(_target_row(row, counts.get(_value(row, "id"), 0)) for row in targets),
        filters=filters,
        widths=(10, 12, 12, 28, 20, 18, 24, 10, 16, 20, 12, 36),
        date_columns={10},
        number_columns={1, 9, 11},
        number_formats={1: "0", 9: "#,##0", 11: "#,##0"},
        tab_color="FE2C55",
        freeze_panes="D2",
        row_height=32,
        generated_at=generated_at,
    )


def build_watches_report(
    watches: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    return build_module_report(
        report_title="CreatorHub 评论监控配置报告",
        module_name="评论监控",
        sheet_title="评论监控",
        headers=_WATCH_HEADERS,
        rows=(_watch_row(row) for row in watches),
        filters=filters,
        widths=(10, 12, 14, 30, 20, 18, 24, 14, 10, 12, 16, 20, 36),
        date_columns={12},
        number_columns={1, 10, 11},
        number_formats={1: "0", 10: "#,##0", 11: "#,##0"},
        tab_color="14B8A6",
        freeze_panes="D2",
        row_height=32,
        generated_at=generated_at,
    )


def build_danmaku_watches_report(
    watches: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    return build_module_report(
        report_title="CreatorHub 弹幕监控配置报告",
        module_name="弹幕监控",
        sheet_title="弹幕监控",
        headers=_DANMAKU_WATCH_HEADERS,
        rows=(_danmaku_watch_row(row) for row in watches),
        filters=filters,
        widths=(10, 12, 14, 30, 20, 18, 24, 14, 10, 12, 16, 20, 36),
        date_columns={12},
        number_columns={1, 10, 11},
        number_formats={1: "0", 10: "#,##0", 11: "#,##0"},
        tab_color="8B5CF6",
        freeze_panes="D2",
        row_height=32,
        generated_at=generated_at,
    )


def build_share_history_report(
    records: Sequence[Any],
    *,
    filters: Sequence[tuple[str, Any]] = (),
    generated_at: datetime | None = None,
) -> bytes:
    return build_module_report(
        report_title="CreatorHub 链接下载历史报告",
        module_name="链接下载历史",
        sheet_title="链接下载历史",
        headers=_SHARE_HISTORY_HEADERS,
        rows=(_share_history_row(row) for row in records),
        filters=filters,
        widths=(10, 12, 24, 48, 22, 14, 12, 20, 20, 12, 12, 12, 14, 14, 40, 50, 50, 50),
        date_columns={8, 9},
        number_columns={1, 7, 10, 11, 12},
        number_formats={1: "0", 7: "#,##0", 10: "#,##0", 11: "#,##0", 12: "0.0"},
        link_columns={15, 16, 17},
        tab_color="FE2C55",
        freeze_panes="D2",
        row_height=36,
        generated_at=generated_at,
    )


def build_keyword_collection_report(job: Any, contents: Sequence[Any],
                                    comments: Sequence[Any],
                                    generated_at: datetime | None = None) -> bytes:
    """Build one workbook containing the keyword job, content and comments."""
    generated_at = generated_at or datetime.now()
    workbook = Workbook()
    _configure_workbook(workbook)
    workbook.properties.creator = "CreatorHub"
    workbook.properties.title = "CreatorHub 关键词采集报告"
    keyword_value = _value(job, "keywords", "[]")
    try:
        keyword_text = ", ".join(json.loads(keyword_value or "[]"))
    except (TypeError, ValueError, json.JSONDecodeError):
        keyword_text = _text(keyword_value)
    _write_module_summary(
        workbook,
        report_title="CreatorHub 关键词采集报告",
        module_name="关键词采集",
        detail_sheet="作品",
        record_count=len(contents),
        filters=(
            ("任务 ID", _value(job, "id")),
            ("平台", _value(job, "platform")),
            ("关键词", keyword_text),
            ("任务状态", _value(job, "status")),
            ("作品数", len(contents)),
            ("评论数", len(comments)),
        ),
        generated_at=generated_at,
    )
    content_by_id = {_value(row, "id"): row for row in contents}
    content_rows = []
    for row in contents:
        platform = _text(_value(row, "platform"))
        aweme_id = _text(_value(row, "aweme_id"))
        url = (f"https://www.xiaohongshu.com/explore/{aweme_id}"
               if platform == "xhs" else f"https://www.douyin.com/video/{aweme_id}")
        content_rows.append((
            _value(row, "id"), _text(_value(row, "keyword")), platform,
            aweme_id, _text(_value(row, "desc")),
            _text(_value(row, "author_name")), _text(_value(row, "media_type")),
            _timestamp(_value(row, "create_time")), _value(row, "like_count", 0),
            _value(row, "comment_count", 0),
            _value(row, "collected_comment_count", 0),
            _text(_value(row, "download_status")),
            _local_path_text(_value(row, "local_path")), url,
            _text(_value(row, "error")),
        ))
    _write_table(
        workbook, "作品",
        ("ID", "关键词", "平台", "作品ID", "描述", "作者", "媒体类型",
         "发布时间", "点赞数", "平台评论数", "已采评论数", "下载状态",
         "本地文件", "作品链接", "错误"),
        content_rows,
        widths=(10, 20, 12, 24, 54, 22, 12, 20, 12, 14, 14, 14, 42, 42, 38),
        date_columns={8}, number_columns={1, 9, 10, 11},
        link_columns={13, 14}, tab_color="E11D48", freeze_panes="E2",
        row_height=36,
    )
    comment_rows = []
    for row in comments:
        content = content_by_id.get(_value(row, "content_id"))
        comment_rows.append((
            _value(row, "id"), _text(_value(content, "keyword")),
            _text(_value(row, "platform")), _text(_value(row, "aweme_id")),
            _text(_value(row, "comment_id")), _text(_value(row, "text")),
            _text(_value(row, "user_nickname")), _value(row, "like_count", 0),
            _timestamp(_value(row, "create_time")),
            "回复" if _value(row, "reply_to") else "一级评论",
            _text(_value(row, "reply_to")),
        ))
    _write_table(
        workbook, "评论",
        ("ID", "关键词", "平台", "作品ID", "评论ID", "评论内容", "用户昵称",
         "点赞数", "评论时间", "类型", "上级评论ID"),
        comment_rows,
        widths=(10, 20, 12, 24, 24, 58, 22, 12, 20, 14, 24),
        date_columns={9}, number_columns={1, 8}, tab_color="14B8A6",
        freeze_panes="F2", row_height=36,
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_summary(
    workbook: Workbook,
    *,
    platform: str,
    period_start: date | None,
    period_end: date | None,
    generated_at: datetime,
    targets: Sequence[Any],
    contents: Sequence[Any],
    comments: Sequence[Any],
    danmaku: Sequence[Any],
) -> None:
    worksheet = workbook.create_sheet("概览", 0)
    _configure_sheet(
        worksheet,
        tab_color="0F766E",
        freeze_panes="A3",
        landscape=False,
    )
    worksheet.sheet_format.defaultRowHeight = 22
    worksheet.merge_cells("A1:F1")
    title = worksheet["A1"]
    title.value = "CreatorHub 监控数据报告"
    title.fill = _TITLE_FILL
    title.font = _TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30
    for column in range(1, 7):
        worksheet.cell(row=1, column=column).fill = _TITLE_FILL

    worksheet.merge_cells("A2:F2")
    worksheet["A2"] = "跨模块采集汇总 · 数据明细保留原始时间、标识与筛选上下文"
    worksheet["A2"].font = Font(name="Microsoft YaHei", size=10, color="0F766E")
    worksheet["A2"].fill = _SUBTITLE_FILL
    worksheet["A2"].alignment = _WRAP

    end_label = period_end.isoformat() if period_end else "不限"
    start_label = period_start.isoformat() if period_start else "不限"
    metadata = [
        ("生成时间", generated_at),
        ("平台", platform or "全部平台"),
        ("统计周期", f"{start_label} 至 {end_label}"),
        ("统计口径", "按记录采集时间筛选；作品/评论/弹幕明细保留原始发布时间"),
    ]
    for row, (key, value) in enumerate(metadata, 3):
        label = worksheet.cell(row=row, column=1, value=key)
        label.font = _LABEL_FONT
        label.fill = _META_LABEL_FILL
        label.alignment = _CENTER
        label.border = _THIN_BORDER
        cell = worksheet.cell(row=row, column=2, value=value)
        cell.font = _BODY_FONT
        cell.alignment = _WRAP
        cell.fill = _META_VALUE_FILL
        cell.border = _THIN_BORDER
        if isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    metrics = [
        ("监控目标", len(targets)),
        ("作品记录", len(contents)),
        ("评论记录", len(comments)),
        ("弹幕记录", len(danmaku)),
        ("已下载作品", sum(_value(row, "download_status") == "done" for row in contents)),
        ("作品点赞总数", sum(int(_value(row, "like_count", 0) or 0) for row in contents)),
        ("评论点赞总数", sum(int(_value(row, "like_count", 0) or 0) for row in comments)),
        ("弹幕点赞总数", sum(int(_value(row, "like_count", 0) or 0) for row in danmaku)),
    ]
    worksheet["A9"] = "核心指标"
    worksheet["A9"].fill = _SUBTITLE_FILL
    worksheet["A9"].font = Font(name="Microsoft YaHei", bold=True, color="0F766E")
    worksheet.merge_cells("A9:B9")
    for column in range(1, 3):
        worksheet.cell(row=9, column=column).fill = _SUBTITLE_FILL
    for row, (key, value) in enumerate(metrics, 10):
        label = worksheet.cell(row=row, column=1, value=key)
        label.font = _LABEL_FONT
        label.fill = _META_LABEL_FILL
        label.border = _THIN_BORDER
        value_cell = worksheet.cell(row=row, column=2, value=value)
        value_cell.font = Font(name="Microsoft YaHei", bold=True, size=12, color="0F766E")
        value_cell.fill = _KPI_FILL
        value_cell.border = _THIN_BORDER
        value_cell.alignment = _NUMBER
        value_cell.number_format = "#,##0"

    worksheet["D9"] = "数据集"
    worksheet["D9"].fill = _SUBTITLE_FILL
    worksheet["D9"].font = Font(name="Microsoft YaHei", bold=True, color="0F766E")
    worksheet.merge_cells("D9:F9")
    for column in range(4, 7):
        worksheet.cell(row=9, column=column).fill = _SUBTITLE_FILL
    dataset_rows = [
        ("监控目标", "监控目标"),
        ("作品数据", "作品数据"),
        ("评论数据", "评论数据"),
        ("弹幕数据", "弹幕数据"),
    ]
    for row, (key, sheet_name) in enumerate(dataset_rows, 10):
        dataset_label = worksheet.cell(row=row, column=4, value=key)
        dataset_label.font = _LABEL_FONT
        dataset_label.fill = _META_LABEL_FILL
        dataset_label.border = _THIN_BORDER
        count_cell = worksheet.cell(row=row, column=5, value=len({
            "监控目标": targets,
            "作品数据": contents,
            "评论数据": comments,
            "弹幕数据": danmaku,
        }[sheet_name]))
        count_cell.font = _BODY_FONT
        count_cell.fill = _KPI_FILL
        count_cell.border = _THIN_BORDER
        count_cell.alignment = _NUMBER
        count_cell.number_format = "#,##0"
        detail_cell = worksheet.cell(row=row, column=6, value=f"见《{sheet_name}》工作表")
        detail_cell.font = _LINK_FONT
        detail_cell.fill = _LINK_FILL
        detail_cell.border = _THIN_BORDER
        _internal_hyperlink(detail_cell, sheet_name)

    daily = Counter()
    for collection in (contents, comments, danmaku):
        for row in collection:
            value = _excel_datetime(_value(row, "created_at"))
            if value:
                daily[value.date().isoformat()] += 1
    worksheet["A20"] = "按采集日期汇总"
    worksheet["A20"].fill = _SUBTITLE_FILL
    worksheet["A20"].font = Font(name="Microsoft YaHei", bold=True, color="0F766E")
    worksheet.merge_cells("A20:C20")
    for column in range(1, 4):
        worksheet.cell(row=20, column=column).fill = _SUBTITLE_FILL
    for col, header in enumerate(("日期", "新增记录", "说明"), 1):
        cell = worksheet.cell(row=21, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    for row, day in enumerate(sorted(daily, reverse=True)[:31], 22):
        day_cell = worksheet.cell(row=row, column=1, value=datetime.strptime(day, "%Y-%m-%d").date())
        day_cell.font = _BODY_FONT
        day_cell.number_format = "yyyy-mm-dd"
        day_cell.border = _THIN_BORDER
        count_cell = worksheet.cell(row=row, column=2, value=daily[day])
        count_cell.font = _BODY_FONT
        count_cell.alignment = _NUMBER
        count_cell.number_format = "#,##0"
        count_cell.border = _THIN_BORDER
        detail_cell = worksheet.cell(row=row, column=3, value="作品 + 评论 + 弹幕")
        detail_cell.font = _BODY_FONT
        detail_cell.border = _THIN_BORDER
        if row % 2 == 0:
            for column in range(1, 4):
                worksheet.cell(row=row, column=column).fill = _ALT_ROW_FILL

    for column, width in {"A": 22, "B": 24, "C": 34, "D": 18, "E": 14, "F": 26}.items():
        worksheet.column_dimensions[column].width = width
    last_row = max(21, 21 + min(len(daily), 31))
    worksheet.print_area = f"A1:F{last_row}"
    worksheet.oddFooter.center.text = "&A  ·  第 &P / &N 页"


def build_monitor_report(
    *,
    platform: str,
    period_start: date | None,
    period_end: date | None,
    targets: Sequence[Any],
    contents: Sequence[Any],
    watches: Sequence[Any],
    comments: Sequence[Any],
    danmaku_watches: Sequence[Any],
    danmaku: Sequence[Any],
    generated_at: datetime | None = None,
) -> bytes:
    """Build a formatted workbook from already-filtered monitoring records."""
    workbook = Workbook()
    _configure_workbook(workbook)
    workbook.remove(workbook.active)
    generated_at = generated_at or datetime.now()
    workbook.properties.creator = "CreatorHub"
    workbook.properties.title = "CreatorHub 监控数据报告"
    workbook.properties.subject = "跨模块监控数据 Excel 导出"

    _write_summary(
        workbook,
        platform=platform,
        period_start=period_start,
        period_end=period_end,
        generated_at=generated_at,
        targets=targets,
        contents=contents,
        comments=comments,
        danmaku=danmaku,
    )

    target_names = _target_map(targets)
    watch_names = _watch_map(watches)
    danmaku_watch_names = _watch_map(danmaku_watches)

    _write_table(
        workbook,
        "监控目标",
        ("ID", "平台", "类型", "目标名称", "别名", "分组", "标签", "启用", "检查间隔(秒)", "最后扫描", "作品数", "最近错误"),
        (
            (
                _value(row, "id"),
                _text(_value(row, "platform")),
                _text(_value(row, "target_kind")),
                _label(row),
                _text(_value(row, "alias")),
                _text(_value(row, "group_name")),
                _tags(_value(row, "tags")),
                "是" if _value(row, "enabled", True) else "否",
                _value(row, "interval_seconds", 0),
                _excel_datetime(_value(row, "last_scan_at")),
                sum(1 for item in contents if _value(item, "target_id") == _value(row, "id")),
                _text(_value(row, "last_error")),
            )
            for row in targets
        ),
        widths=(10, 12, 12, 28, 20, 18, 24, 10, 16, 20, 12, 36),
        date_columns={10},
        number_columns={1, 9, 11},
        number_formats={1: "0", 9: "#,##0", 11: "#,##0"},
        tab_color="FE2C55",
        freeze_panes="D2",
        row_height=32,
    )

    _write_table(
        workbook,
        "作品数据",
        ("ID", "监控目标", "平台", "作品ID", "描述", "媒体类型", "发布时间", "采集时间", "点赞数", "评论数", "时长(秒)", "画质", "下载状态", "本地文件", "封面链接", "错误"),
        (
            (
                _value(row, "id"),
                target_names.get(_value(row, "target_id"), f"目标#{_value(row, 'target_id')}"),
                _text(_value(row, "platform")),
                _text(_value(row, "aweme_id")),
                _text(_value(row, "desc")),
                _text(_value(row, "media_type")),
                _timestamp(_value(row, "create_time")),
                _excel_datetime(_value(row, "created_at")),
                _value(row, "like_count", 0),
                _value(row, "comment_count", 0),
                _value(row, "duration", 0),
                _text(_value(row, "quality")),
                _text(_value(row, "download_status")),
        _local_path_text(_value(row, "local_path")),
                _text(_value(row, "cover_url")),
                _text(_value(row, "error")),
            )
            for row in contents
        ),
        widths=(10, 28, 12, 24, 50, 12, 20, 20, 12, 12, 12, 12, 14, 40, 40, 36),
        date_columns={7, 8},
        number_columns={1, 9, 10, 11},
        number_formats={1: "0", 9: "#,##0", 10: "#,##0", 11: "0.0"},
        link_columns={14, 15},
        tab_color="2563EB",
        freeze_panes="E2",
        row_height=34,
    )

    _write_table(
        workbook,
        "评论数据",
        ("ID", "评论监控", "平台", "作品ID", "评论ID", "评论内容", "用户昵称", "点赞数", "评论时间", "采集时间", "类型", "上级评论ID"),
        (
            (
                _value(row, "id"),
                watch_names.get(_value(row, "watch_id"), f"监控#{_value(row, 'watch_id')}"),
                _text(_value(row, "platform")),
                _text(_value(row, "aweme_id")),
                _text(_value(row, "comment_id")),
                _text(_value(row, "text")),
                _text(_value(row, "user_nickname")),
                _value(row, "like_count", 0),
                _timestamp(_value(row, "create_time")),
                _excel_datetime(_value(row, "created_at")),
                "回复" if _value(row, "reply_to") else "一级评论",
                _text(_value(row, "reply_to")),
            )
            for row in comments
        ),
        widths=(10, 28, 12, 24, 24, 56, 22, 12, 20, 20, 14, 24),
        date_columns={9, 10},
        number_columns={1, 8},
        number_formats={1: "0", 8: "#,##0"},
        tab_color="14B8A6",
        freeze_panes="F2",
        row_height=36,
    )

    _write_table(
        workbook,
        "弹幕数据",
        ("ID", "弹幕监控", "平台", "作品ID", "弹幕ID", "弹幕内容", "用户ID", "用户昵称", "视频内时间(秒)", "发送时间", "采集时间", "点赞数", "来源", "已屏蔽"),
        (
            (
                _value(row, "id"),
                danmaku_watch_names.get(_value(row, "watch_id"), f"监控#{_value(row, 'watch_id')}"),
                _text(_value(row, "platform")),
                _text(_value(row, "aweme_id")),
                _text(_value(row, "danmaku_id")),
                _text(_value(row, "text")),
                _text(_value(row, "user_id")),
                _text(_value(row, "user_nickname")),
                round(int(_value(row, "video_time_ms", 0) or 0) / 1000, 3),
                _timestamp(_value(row, "create_time")),
                _excel_datetime(_value(row, "created_at")),
                _value(row, "like_count", 0),
                _text(_value(row, "source")),
                "是" if _value(row, "is_blocked", False) else "否",
            )
            for row in danmaku
        ),
        widths=(10, 28, 12, 24, 24, 56, 24, 22, 18, 20, 20, 12, 12, 12),
        date_columns={10, 11},
        number_columns={1, 9, 12},
        number_formats={1: "0", 9: "0.000", 12: "#,##0"},
        tab_color="8B5CF6",
        freeze_panes="F2",
        row_height=36,
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
